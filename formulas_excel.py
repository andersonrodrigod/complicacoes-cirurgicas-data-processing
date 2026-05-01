from openpyxl import load_workbook


FORMULAS_BASE = [
    {
        "coluna": "DT ENVIO",
        "letra": "AB",
        "formula": lambda linha: (
            f'=SEERRO(SE(AI{linha}<>"",AI{linha},'
            f'ÍNDICE(STATUS!R:R,CORRESP(AF{linha},STATUS!O:O,0))),"")'
        ),
    },
    {
        "coluna": "CHAVE",
        "letra": "AF",
        "formula": lambda linha: f'=J{linha}&"_"&V{linha}&"_"&Y{linha}&"_"&AA{linha}&"_"&F{linha}',
    },
    {
        "coluna": "STATUS",
        "letra": "AM",
        "formula": lambda linha: (
            f'=SE(OU(AK{linha}="Não quis",AK{linha}="Óbito",AK{linha}="Lida"),'
            f'AK{linha},SE(CONT.SES(STATUS!O:O,AF{linha},STATUS!G:G,"Lida")>0,"Lida",""))'
        ),
    },
    {
        "coluna": "P1",
        "letra": "AO",
        "formula": lambda linha: f'=PROCX(AF{linha},\'P1\'!B:B,\'P1\'!E:E,"")',
    },
    {
        "coluna": "P2",
        "letra": "AP",
        "formula": lambda linha: f'=PROCX(AF{linha},\'P2\'!B:B,\'P2\'!E:E,"")',
    },
    {
        "coluna": "P3",
        "letra": "AQ",
        "formula": lambda linha: f'=PROCX(AF{linha},\'P3\'!B:B,\'P3\'!E:E,"")',
    },
    {
        "coluna": "P4",
        "letra": "AR",
        "formula": lambda linha: f'=PROCX(AF{linha},\'P4\'!B:B,\'P4\'!E:E,"")',
    },
]


def ultima_linha_com_dados(ws, coluna_referencia="J"):
    for linha in range(ws.max_row, 1, -1):
        valor = ws[f"{coluna_referencia}{linha}"].value
        if valor not in (None, ""):
            return linha
    return 1


def validar_posicoes_colunas(ws):
    erros = []

    for regra in FORMULAS_BASE:
        valor_cabecalho = ws[f"{regra['letra']}1"].value

        if valor_cabecalho == regra["coluna"]:
            print(f"OK formula: {regra['coluna']} encontrada em {regra['letra']}")
        else:
            mensagem = (
                f"ERRO formula: esperado {regra['coluna']} em {regra['letra']}, "
                f"mas encontrado {valor_cabecalho}"
            )
            print(mensagem)
            erros.append(mensagem)

    if erros:
        raise ValueError("Posicionamento de colunas invalido para aplicar formulas.")


def aplicar_formulas(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb["BASE"]

    validar_posicoes_colunas(ws)
    ultima_linha = ultima_linha_com_dados(ws)

    if ultima_linha < 2:
        print("Nenhuma linha de dados encontrada para aplicar formulas.")
        wb.save(caminho_arquivo)
        return

    for regra in FORMULAS_BASE:
        for linha in range(2, ultima_linha + 1):
            ws[f"{regra['letra']}{linha}"] = regra["formula"](linha)

        print(f"Formula aplicada em {regra['coluna']} ({regra['letra']}2:{regra['letra']}{ultima_linha})")

    wb.save(caminho_arquivo)
